import openpyxl
from copy import copy
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from dateutil import parser

def clear_current_report(current_dir,filename):
    """Очищает текущий Отчёт по ОС"""

    Report = openpyxl.load_workbook(current_dir+filename)  # Our Everyday Report
    counting1 = Report['Подсчет1']
    index=2
    while True:
        if counting1.cell(row=index, column=2).value != None:
            for i in range(2,9):
                counting1.cell(row=index, column=i).value = None
        else:
            break
        index += 1

    table_for_count2 = Report['Таблица для Подсчета2'] #table_for_count_alm.cell(row=2, column=1)
    index = 2                                          #table_for_count2.cell(row=2, column=1)
    while True:
        if table_for_count2.cell(row=index, column=1).value != None:
            for i in range(1,12):
                table_for_count2.cell(row=index, column=i).value = None
        else:
            break
        index+=1

    table_for_count_alm = Report['Таблица для Посчета ALM']
    index = 2
    while True:
        if table_for_count_alm.cell(row=index, column=1).value != None:
            for i in range(1,13):
                table_for_count_alm.cell(row=index, column=i).value = None

                # new_cell.font = copy(cell.font)
                # new_cell.border = copy(cell.border)
                # new_cell.fill = copy(cell.fill)
                # new_cell.number_format = copy(cell.number_format)
                # new_cell.protection = copy(cell.protection)
                # new_cell.alignment = copy(cell.alignment)
                table_for_count_alm.cell(row=index, column=i).font = copy(table_for_count_alm.cell(row=2, column=16).font)
                table_for_count_alm.cell(row=index, column=i).border = copy(table_for_count_alm.cell(row=2, column=16).border)
                table_for_count_alm.cell(row=index, column=i).fill = copy(table_for_count_alm.cell(row=2, column=16).fill)

        else:
            break
        index+=1

    omni = Report['Таблица для Подсчета Omni']
    index = 2
    while True:
        if omni.cell(row=index, column=1).value != None:
            for i in range(1, 12):
                omni.cell(row=index, column=i).value = None
                omni.cell(row=index, column=i).font = copy(omni.cell(row=2, column=16).font)
                omni.cell(row=index, column=i).border = copy(omni.cell(row=2, column=16).border)
                omni.cell(row=index, column=i).fill = copy(omni.cell(row=2, column=16).fill)
        else:
            break
        index += 1

    Report.save(current_dir+filename)  # E:\\_proj\\Neoflex\\_everyday\\Report.xlsx


def import_from_google_tables_oscontrol(curdate):
    """Import Data from Google Tables"""
    omni_key,omni_sheet_name = '1drbPbjMKGbODn1FGqmR0VdRzh3zyaxGVjq1prTAu2rs','2019'
    oscontrol_key,oscontrol_sheet_name='1HiDdPqB_-ro4Iu0RplDnt3k8lEjDd_UiOBPKvyWOPuY','Сводная таблица по OS'
    almcontrol_key,almcontrol_sheet_name='1SDSEhgtQTHR9a69BfYE1Fd2DnMH6fIFBEsvyHldwmM8','Сводная таблица по ALM'####for test '13criem2KpgGtQA3fjv2BxH3f3-r3WWi5TJqod4bAQfg','Лист1'

    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('NeoflexReports-376d91d0718a.json', scope)
    client = gspread.authorize(creds)
    
    oscontrol_sheet = client.open_by_key(oscontrol_key).worksheet(oscontrol_sheet_name).get_all_values()  ##Should be *.worksheet('OSCONTROL')
    oscontrol_list = []  # list with OS with all filled cells
    for i in range(len(oscontrol_sheet), 0, -1):
        try:
            if parser.parse(oscontrol_sheet[i - 1][2], dayfirst=True).date() == curdate:
                flag1 = 1
                if oscontrol_sheet[i - 1][0] != '' and oscontrol_sheet[i - 1][1] != '' and oscontrol_sheet[i - 1][2] != '' and oscontrol_sheet[i - 1][11] != '':
                    oscontrol_list.append(oscontrol_sheet[i - 1][0:12])
                else:
                    continue
            else:
                for x in range(2,5):
                    if parser.parse(oscontrol_sheet[i - x][2], dayfirst=True).date() == curdate:
                        flag1=0
                        break
                try:
                    if flag1==1:
                        break
                except Exception as e:
                    continue     
        except Exception as e:
            #print(str(e),oscontrol_sheet[i - 1],i)
            continue 


    omni_sheet = client.open_by_key(omni_key).worksheet(omni_sheet_name).get_all_values()
    omni_list=[]
    for i in range(len(omni_sheet), 0, -1):
        try:
            if parser.parse(omni_sheet[i - 1][2], dayfirst=True).date() == curdate:
                flag2 = 1
                if omni_sheet[i - 1][5] != '' and omni_sheet[i - 1][0] != '' and omni_sheet[i - 1][1] != '' and omni_sheet[i - 1][2] != '':
                    omni_list.append(omni_sheet[i - 1][0:2]+omni_sheet[i - 1][3:6]+omni_sheet[i - 1][7:])
                else:
                    continue
            else:      
                for x in range(2,5):
                    if parser.parse(omni_sheet[i - x][2], dayfirst=True).date() == curdate:
                        flag2=0
                        break
                try:
                    if flag2==1:
                        break
                except Exception as e:
                    continue   
        except Exception as e:
            #print(str(e),omni_sheet[i - 1],i)
            continue       

    alm_sheet= client.open_by_key(almcontrol_key).worksheet(almcontrol_sheet_name).get_all_values()
    alm_list=[]
    for i in range(len(alm_sheet), 0, -1):
        try:
            if parser.parse(alm_sheet[i - 1][2], dayfirst=True).date() == curdate:
                flag3=1
                if alm_sheet[i - 1][0] != '' and alm_sheet[i - 1][1] != '' and alm_sheet[i - 1][2] != '' and alm_sheet[i - 1][11] != '':
                    alm_list.append(alm_sheet[i - 1][0:12])
                else:
                    continue
            else:
                for x in range(2,5):
                    if parser.parse(alm_sheet[i - x][2], dayfirst=True).date() == curdate:
                        flag3=0
                        break
                try:
                    if flag3==1:
                        break
                except Exception as e:
                    continue          
        except Exception as e:
            #print(str(e),alm_sheet[i - 1],i)
            continue


    return oscontrol_list,omni_list,alm_list
